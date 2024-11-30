import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import calendar

# Configuração da janela principal
window = tk.Tk()
window.title("App de Poupança Pessoal")
window.geometry("700x500")
window.configure(bg="#252525")

# Configuração de estilos
style = ttk.Style()
style.theme_use("clam")
style.configure("TLabel",
                background="#252525",
                foreground="#FFFFFF",
                font=("Arial", 12))
style.configure("TEntry",
                fieldbackground="#FFFFFF",
                font=("Arial", 12))
style.configure("TButton",
                background="#4CAF50",
                foreground="#FFFFFF",
                font=("Arial", 12))

# Variáveis globais
valores = []
canvas = None

# Função para salvar o valor
def salvar_valor():
    global canvas

    try:
        valor_dia = float(entry_valor.get())
        valores.append(valor_dia)
        total_valores = sum(valores)

        entry_valor.delete(0, tk.END)
        label_status.config(text="Valor salvo com sucesso!", foreground="green")
        label_total.config(text=f"Total economizado: R${total_valores:.2f}")

        # Atualizar Excel (Simulação)
        linha = len(valores) + 1
        coluna_data = get_column_letter(1)
        coluna_valor = get_column_letter(2)

        # Atualização da planilha seria feita aqui
        # Simulação:
        datas.append(date.today())
        valores_salvos.append(valor_dia)

        # Atualizar gráfico
        if canvas:
            canvas.get_tk_widget().destroy()
        plotar_grafico()

    except ValueError:
        label_status.config(text="Por favor, insira um valor numérico válido.", foreground="red")

# Função para plotar o gráfico
def plotar_grafico():
    global canvas

    fig = plt.Figure(figsize=(12, 6), dpi=80)
    ax_barras = fig.add_subplot(121)
    ax_pie = fig.add_subplot(122)

    # Processar dados mensais
    dados_mensais = {}
    for data, valor in zip(datas, valores_salvos):
        mes_ano = data.strftime("%m-%Y")
        if mes_ano in dados_mensais:
            dados_mensais[mes_ano].append(valor)
        else:
            dados_mensais[mes_ano] = [valor]

    # Gráfico de barras
    barras = ax_barras.bar(range(len(dados_mensais)), [sum(v) for v in dados_mensais.values()])
    nomes_meses = [calendar.month_name[int(m.split('-')[0])] + '-' + m.split('-')[1] for m in dados_mensais.keys()]

    ax_barras.set_xticks(range(len(nomes_meses)))
    ax_barras.set_xticklabels(nomes_meses, rotation=45, ha='right')
    ax_barras.set_title('Economia por Mês')

    # Gráfico de pizza
    valores_por_semana = [sum(valor for data, valor in zip(datas, valores_salvos)
                              if data.weekday() < 7) for i in range(4)]  # Ajustar semanas
    ax_pie.pie(valores_por_semana, labels=[f'Semana {i+1}' for i in range(4)],
               autopct='%1.1f%%', startangle=90)
    ax_pie.set_title('Economia por Semana')

    # Inserir gráficos no Tkinter
    canvas = FigureCanvasTkAgg(fig, master=window)
    canvas.get_tk_widget().pack(pady=10)

# Elementos da interface
label_instrucao = ttk.Label(window, text="Insira o valor diário:")
label_status = ttk.Label(window, text="", foreground="red")
label_total = ttk.Label(window, text="", font=("Arial", 14, "bold"))

entry_valor = ttk.Entry(window)
button_salvar = ttk.Button(window, text="Salvar", command=salvar_valor)

# Posicionamento dos elementos
label_instrucao.pack(pady=10)
entry_valor.pack(pady=5)
button_salvar.pack(pady=10)
label_status.pack()
label_total.pack(pady=10)

# Variáveis para armazenar dados
datas = []
valores_salvos = []

# Iniciar a janela
window.mainloop()
